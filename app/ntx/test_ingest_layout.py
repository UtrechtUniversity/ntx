from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from .ingest.layout import LayoutError, parse_layout_xlsx
from .ingest.wells import parse_well_string


def test_parse_well_string_range_and_single():
    wells = parse_well_string("A1-A4 B2 C3-C4")
    assert wells == ["A1", "A2", "A3", "A4", "B2", "C3", "C4"]


def test_parse_well_string_2d_range():
    wells = parse_well_string("A1-B2")
    assert wells == ["A1", "A2", "B1", "B2"]


def test_parse_well_string_invalid():
    with pytest.raises(ValueError):
        parse_well_string("A0")
    with pytest.raises(ValueError):
        parse_well_string("B2-A1")


def test_parse_real_layout_fixture(data_dir: Path):
    layout = parse_layout_xlsx(
        data_dir / "201210_LvM_256135_1294-67_LO.xlsx",
    )

    assert layout.date == date(2020, 10, 12)
    assert layout.plate_wells == 48
    assert len(layout.conditions) == 6

    control_conditions = [cond for cond in layout.conditions if cond.is_control]
    assert len(control_conditions) == 1
    control = control_conditions[0]
    assert control.is_control
    assert control.concentration is None
    assert {"B1", "B2", "B3"}.issubset(set(control.wells))

    high = next(cond for cond in layout.conditions if cond.concentration == Decimal("100"))
    assert not high.is_control
    assert "A1" in high.wells

    all_wells = [well for condition in layout.conditions for well in condition.wells]
    assert len(all_wells) == layout.plate_wells
    assert len(set(all_wells)) == len(all_wells)


def test_missing_date_raises(tmp_path: Path):
    tmp_path = tmp_path / "missing_date.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        pytest.fail("Workbook has no active worksheet")
    ws.append(["Wells", 48])
    ws.append(["Groups", None])
    ws.append(["Control", "A1"])
    wb.save(tmp_path)

    with pytest.raises(LayoutError):
        parse_layout_xlsx(tmp_path)


def test_parse_layout_date_dd_mm_yyyy(tmp_path: Path):
    tmp_path = tmp_path / "dd_mm_yyyy.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        pytest.fail("Workbook has no active worksheet")
    ws.append(["Date", "28/09/2023"])
    ws.append(["Wells", 48])
    ws.append(["Groups", None])
    ws.append(["Control", "A1"])
    ws.append(["0.1", "A2"])
    wb.save(tmp_path)

    layout = parse_layout_xlsx(tmp_path)
    assert layout.date == date(2023, 9, 28)


def test_duplicate_wells_raise(tmp_path: Path):
    tmp_path = tmp_path / "duplicate_wells.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        pytest.fail("Workbook has no active worksheet")
    ws.append(["Date", date(2024, 1, 1)])
    ws.append(["Wells", 48])
    ws.append(["Groups", None])
    ws.append(["Control", "A1 A2"])
    ws.append(["0.1", "A2 A3"])
    wb.save(tmp_path)

    with pytest.raises(LayoutError):
        parse_layout_xlsx(tmp_path)


def test_wells_outside_plate_raise(tmp_path: Path):
    tmp_path = tmp_path / "out_of_range_wells.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        pytest.fail("Workbook has no active worksheet")
    ws.append(["Date", date(2024, 1, 1)])
    ws.append(["Wells", 48])
    ws.append(["Groups", None])
    ws.append(["Control", "A1"])
    ws.append(["0.1", "H12"])
    wb.save(tmp_path)

    with pytest.raises(LayoutError):
        parse_layout_xlsx(tmp_path)
