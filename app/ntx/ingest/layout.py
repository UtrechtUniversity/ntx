"""Parsing of Axion layout workbooks + legacy group parsing."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .wells import parse_well_string

logger = logging.getLogger(__name__)

PLATE_DIMENSIONS: dict[int, tuple[int, int]] = {
    48: (6, 8),
}


class LayoutError(ValueError):
    """Raised when a layout workbook cannot be parsed."""


@dataclass
class ConditionLayout:
    concentration: Decimal | None
    wells: list[str]
    is_control: bool
    chemical: str | None = None
    unit: str | None = None 


@dataclass
class ExperimentLayout:
    date: date
    plate_wells: int
    conditions: list[ConditionLayout]


def parse_group_name(name: str) -> dict[str, object]:
    name = str(name).strip()
    parts = name.split()
    chemical, concentration, unit = None, None, None

    pattern = re.compile(r"([\d\.]+)\s*([a-zA-Z/%µμ]*)")

    for i, part in enumerate(parts):
        match = pattern.fullmatch(part)
        if match:
            try:
                concentration = float(match.group(1))
            except ValueError:
                concentration = None

            unit = match.group(2)
            if not unit and i + 1 < len(parts):
                unit = parts[i + 1]

            chemical = " ".join(parts[:i]) if i > 0 else None
            break

    if concentration is None:
        for i, part in enumerate(reversed(parts)):
            if part.replace(".", "", 1).isdigit():
                concentration = float(part)
                chemical = " ".join(parts[:-(i + 1)]) if (i + 1) < len(parts) else None
                break

    if chemical is None and concentration is None:
        chemical = name

    return {"chemical": chemical or None, "concentration": concentration, "unit": unit}


def parse_layout_xlsx(path: str | Path) -> ExperimentLayout:
    """
    Parse an Axion layout workbook (``*_LO.xlsx``).

    Sheet layout (first worksheet):
    - Column A: keys ("Date", "Wells", "Groups", then group labels).
    - Column B: values (date, plate size, and well strings).
    - Conditions start after the "Groups" row.
    """
    path = Path(path)
    if not path.exists():
        raise LayoutError(f"Layout file does not exist: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb.worksheets[0]

    experiment_date: date | None = None
    plate_wells: int | None = None
    conditions: list[ConditionLayout] = []
    in_groups = False

    for key_cell, value_cell in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(key_cell).strip() if key_cell is not None else ""
        value = value_cell
        if not key:
            continue

        lowered = key.lower()
        if lowered == "date":
            experiment_date = _parse_date(value, path)
        elif lowered == "wells":
            plate_wells = _parse_plate_wells(value)
        elif lowered == "groups":
            in_groups = True
        elif in_groups:
            wells = parse_well_string(str(value).strip()) if value is not None else []
            if not wells:
                raise LayoutError(f"Condition '{key}' has no wells listed")

            legacy = parse_group_name(key)
            concentration = legacy["concentration"]
            chemical = legacy["chemical"]
            unit = legacy["unit"]

            is_control = any(k in lowered for k in ("control", "dmso"))

            if concentration is None:
                concentration: Decimal | None = None
            else:
                try:
                    concentration = Decimal(str(concentration))
                except Exception as exc:
                    raise LayoutError(f"Invalid concentration value '{key}'") from exc

            conditions.append(
                ConditionLayout(
                    concentration=concentration,
                    wells=wells,
                    is_control=is_control,
                    chemical=chemical,
                    unit=unit,
                )
            )

    if experiment_date is None:
        raise LayoutError("Layout is missing a Date entry")
    if plate_wells is None:
        raise LayoutError("Layout is missing a Wells entry")
    if not conditions:
        raise LayoutError("Layout contains no conditions after the Groups row")
    if not any(cond.is_control for cond in conditions):
        raise LayoutError("Layout must include at least one Control group")

    _validate_wells(conditions, plate_wells)

    return ExperimentLayout(
        date=experiment_date,
        plate_wells=plate_wells,
        conditions=conditions,
    )

def _parse_date(value, path: Path) -> date:
    if value is None:
        raise LayoutError("Date cell is empty")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(value))).date()

    text = str(value).strip()
    parsed = _try_parse_date_text(text)
    if parsed:
        return parsed

    fallback = _parse_date_from_filename(path)
    if fallback:
        return fallback

    raise LayoutError(f"Unrecognized date format '{value}'")


def _try_parse_date_text(text: str) -> date | None:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _parse_date_from_filename(path: Path) -> date | None:
    stem = path.stem
    for token in re.split(r"[\s_]+", stem):
        if not token:
            continue
        if token.isdigit():
            if len(token) == 8:
                try:
                    return datetime.strptime(token, "%Y%m%d").date()
                except ValueError:
                    continue
            if len(token) == 6:
                try:
                    return datetime.strptime(token, "%y%m%d").date()
                except ValueError:
                    continue
        parsed = _try_parse_date_text(token)
        if parsed:
            return parsed
    return None


def _parse_plate_wells(value) -> int:
    try:
        wells = int(value)
    except Exception:
        raise LayoutError(f"Invalid plate well count '{value}'")
    if wells <= 0:
        raise LayoutError("Plate well count must be positive")
    return wells


def _validate_wells(conditions: Iterable[ConditionLayout], plate_wells: int) -> None:
    seen: set[str] = set()
    dimensions = PLATE_DIMENSIONS.get(plate_wells)
    row_limit, col_limit = dimensions if dimensions else (None, None)

    for condition in conditions:
        for well in condition.wells:
            if well in seen:
                raise LayoutError(f"Duplicate well '{well}' in layout")
            seen.add(well)

            row = well[0].upper()
            try:
                col = int(well[1:])
            except ValueError:
                raise LayoutError(f"Invalid well identifier '{well}'")

            if col <= 0:
                raise LayoutError(f"Well number must be positive: '{well}'")

            if row_limit is not None:
                r = ord(row) - ord("A") + 1
                if r > row_limit:
                    raise LayoutError(f"Well '{well}' exceeds row limit")

            if col_limit is not None and col > col_limit:
                raise LayoutError(f"Well '{well}' exceeds column limit")

    if dimensions:
        max_supported = dimensions[0] * dimensions[1]
        if len(seen) > max_supported:
            raise LayoutError(
                f"Layout lists {len(seen)} wells exceeding plate capacity {max_supported}"
            )
    elif len(seen) > plate_wells:
        raise LayoutError(
            f"Layout lists {len(seen)} wells exceeding plate capacity {plate_wells}"
        )
