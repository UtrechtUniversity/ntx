"""Read Axion-derived BMA Excel artifacts for validation/testing."""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from ntx.metrics_metadata import AXION_METRICS_MAP

from .parsers import parse_axion_rows, translate_metrics


@dataclass(frozen=True)
class BMAMetrics:
    wells: list[str]
    baseline: dict[str, dict[str, float | int | None]]
    exposure: dict[str, dict[str, float | int | None]]


@dataclass(frozen=True)
class BMAConditionRatios:
    wells: list[str]
    ratios: dict[str, dict[str, float | int | None]]


@dataclass(frozen=True)
class BMAConditionNormalizedToControl:
    wells: list[str]
    normalized: dict[str, dict[str, float | int | None]]
    mean: dict[str, float | int | None]
    stdev: dict[str, float | int | None]


_CONDITION_RATIO_HEADER_ROW = 3
_CONDITION_RATIO_FIRST_DATA_ROW = 4
# Condition worksheets store per-well ratios (percent of baseline) in columns R-Y.
_CONDITION_RATIO_MIN_COL = 18  # R
_CONDITION_RATIO_MAX_COL = 25  # Y

_CONDITION_NORM_HEADER_ROW = 3
_CONDITION_NORM_FIRST_DATA_ROW = 4
# Condition worksheets store "Normalized to Control" (percent) in columns Z-AG,
# mean in AH, stdev in AI.
_CONDITION_NORM_MIN_COL = 26  # Z
_CONDITION_NORM_MAX_COL = 35  # AI


def load_bma_metrics(path: str | Path) -> BMAMetrics:
    """
    Load per-well baseline/exposure metrics from a BMA workbook.

    The workbook has two worksheets called "baseline" and "exposure" whose
    "Well Averages" sections mirror Axion CSV exports.
    """
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        baseline_ws = workbook["baseline"]
        exposure_ws = workbook["exposure"]
    except KeyError as exc:
        raise ValueError(
            f"BMA workbook {path.name} must contain 'baseline' and 'exposure' worksheets"
        ) from exc

    baseline_rows = _read_well_average_rows(baseline_ws)
    exposure_rows = _read_well_average_rows(exposure_ws)

    baseline_csv = parse_axion_rows(baseline_rows, source=f"{path.name}:baseline")
    exposure_csv = parse_axion_rows(exposure_rows, source=f"{path.name}:exposure")

    if baseline_csv.wells != exposure_csv.wells:
        raise ValueError(
            "BMA baseline/exposure well headers differ. "
            f"baseline_wells={len(baseline_csv.wells)} exposure_wells={len(exposure_csv.wells)}"
        )

    baseline_map = translate_metrics(baseline_csv)
    exposure_map = translate_metrics(exposure_csv)

    return BMAMetrics(wells=baseline_csv.wells, baseline=baseline_map, exposure=exposure_map)


def load_bma_condition_ratios(path: str | Path, sheet_name: str) -> BMAConditionRatios:
    """
    Load per-well ratio metrics from a condition worksheet in a BMA workbook.

    Condition worksheets contain a baseline/exposure/ratio table. Ratios are stored as
    percentages relative to baseline; this helper converts them back to raw ratios.
    """
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as exc:
            raise ValueError(
                f"BMA workbook {path.name} is missing worksheet '{sheet_name}'"
            ) from exc

        wells = _read_condition_ratio_wells(worksheet, source=f"{path.name}:{sheet_name}")
        ratios = _read_condition_ratio_metrics(
            worksheet,
            wells,
            source=f"{path.name}:{sheet_name}",
        )
        return BMAConditionRatios(wells=wells, ratios=ratios)
    finally:
        workbook.close()


def load_bma_condition_normalized_to_control(
    path: str | Path, sheet_name: str
) -> BMAConditionNormalizedToControl:
    """
    Load per-well "Normalized to Control" values from a condition worksheet in a BMA workbook.

    The worksheet stores values as percentages; this helper converts them back to unitless ratios.
    """
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError as exc:
            raise ValueError(
                f"BMA workbook {path.name} is missing worksheet '{sheet_name}'"
            ) from exc

        header = worksheet.cell(row=2, column=_CONDITION_NORM_MIN_COL).value
        if str(header or "").strip().lower() != "normalized to control":
            raise ValueError(
                f"Worksheet {path.name}:{sheet_name} does not contain "
                "a 'Normalized to Control' block"
            )

        wells = _read_condition_normalized_wells(worksheet, source=f"{path.name}:{sheet_name}")
        normalized, mean, stdev = _read_condition_normalized_metrics(
            worksheet,
            wells,
            source=f"{path.name}:{sheet_name}",
        )
        return BMAConditionNormalizedToControl(
            wells=wells,
            normalized=normalized,
            mean=mean,
            stdev=stdev,
        )
    finally:
        workbook.close()


def _read_well_average_rows(worksheet) -> list[list[str]]:
    """
    Convert a worksheet to CSV-like rows; stop before the electrode section.
    """
    rows: list[list[str]] = []
    for raw_row in worksheet.iter_rows(values_only=True):
        if raw_row and isinstance(raw_row[0], str):
            if raw_row[0].strip().lower().startswith("measurement"):
                break

        row = [_cell_to_text(cell) for cell in raw_row]
        while row and row[-1] == "":
            row.pop()
        rows.append(row)
    return rows


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _read_condition_ratio_wells(worksheet, *, source: str) -> list[str]:
    row = next(
        worksheet.iter_rows(
            min_row=_CONDITION_RATIO_HEADER_ROW,
            max_row=_CONDITION_RATIO_HEADER_ROW,
            min_col=_CONDITION_RATIO_MIN_COL,
            max_col=_CONDITION_RATIO_MAX_COL,
            values_only=True,
        )
    )

    wells: list[str] = []
    for value in row:
        if value is None:
            continue
        label = str(value).strip()
        if not label:
            continue
        wells.append(label)

    if not wells:
        raise ValueError(
            f"No well headers found in ratio columns R-Y for condition worksheet {source}"
        )
    return wells


def _read_condition_ratio_metrics(
    worksheet, wells: list[str], *, source: str
) -> dict[str, dict[str, float | int | None]]:
    ratios: dict[str, dict[str, float | int | None]] = {}
    ratio_max_col = _CONDITION_RATIO_MIN_COL + len(wells) - 1

    for row in worksheet.iter_rows(
        min_row=_CONDITION_RATIO_FIRST_DATA_ROW,
        min_col=1,
        max_col=max(ratio_max_col, _CONDITION_RATIO_MIN_COL),
        values_only=True,
    ):
        raw_label = row[0] if row else None
        if raw_label is None:
            continue

        label = str(raw_label).strip()
        if not label:
            continue
        if label.lower().startswith("measurement"):
            break

        internal = AXION_METRICS_MAP.get(label)
        if internal is None:
            continue

        ratio_values = row[_CONDITION_RATIO_MIN_COL - 1 : _CONDITION_RATIO_MIN_COL - 1 + len(wells)]
        ratios[internal] = {
            well: _parse_bma_percent_ratio(value) for well, value in zip(wells, ratio_values)
        }

    if not ratios:
        raise ValueError(f"No ratio metrics found in condition worksheet {source}")
    return ratios


def _parse_bma_percent_ratio(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None

    numeric: float
    if isinstance(value, (int, float)):
        numeric = float(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            numeric = float(text)
        except ValueError:
            return None

    if not math.isfinite(numeric):
        return None
    return numeric / 100.0


def _read_condition_normalized_wells(worksheet, *, source: str) -> list[str]:
    row = next(
        worksheet.iter_rows(
            min_row=_CONDITION_NORM_HEADER_ROW,
            max_row=_CONDITION_NORM_HEADER_ROW,
            min_col=_CONDITION_NORM_MIN_COL,
            max_col=_CONDITION_NORM_MIN_COL + 7,
            values_only=True,
        )
    )

    wells: list[str] = []
    for value in row:
        if value is None:
            continue
        label = str(value).strip()
        if not label:
            continue
        wells.append(label)

    if not wells:
        raise ValueError(
            f"No well headers found in normalized-to-control columns Z-AG for worksheet {source}"
        )
    return wells


def _read_condition_normalized_metrics(
    worksheet, wells: list[str], *, source: str
) -> tuple[
    dict[str, dict[str, float | int | None]],
    dict[str, float | int | None],
    dict[str, float | int | None],
]:
    normalized: dict[str, dict[str, float | int | None]] = {}
    mean: dict[str, float | int | None] = {}
    stdev: dict[str, float | int | None] = {}

    norm_max_col = _CONDITION_NORM_MIN_COL + len(wells) - 1
    for row in worksheet.iter_rows(
        min_row=_CONDITION_NORM_FIRST_DATA_ROW,
        min_col=1,
        max_col=max(_CONDITION_NORM_MAX_COL, norm_max_col),
        values_only=True,
    ):
        raw_label = row[0] if row else None
        if raw_label is None:
            continue

        label = str(raw_label).strip()
        if not label:
            continue
        if label.lower().startswith("measurement"):
            break

        internal = AXION_METRICS_MAP.get(label)
        if internal is None:
            continue

        values = row[_CONDITION_NORM_MIN_COL - 1 : _CONDITION_NORM_MIN_COL - 1 + len(wells)]
        normalized[internal] = {
            well: _parse_bma_percent_ratio(value) for well, value in zip(wells, values)
        }
        mean[internal] = _parse_bma_percent_ratio(row[_CONDITION_NORM_MAX_COL - 2])
        stdev[internal] = _parse_bma_percent_ratio(row[_CONDITION_NORM_MAX_COL - 1])

    if not normalized:
        raise ValueError(f"No normalized-to-control metrics found in worksheet {source}")
    return normalized, mean, stdev
