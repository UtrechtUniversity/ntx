"""Parsing of Axion layout workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .wells import parse_well_string

logger = logging.getLogger(__name__)

# Standard plate dimensions keyed by total wells (rows, columns).
PLATE_DIMENSIONS: dict[int, tuple[int, int]] = {
    48: (6, 8),
}


class LayoutError(ValueError):
    """Raised when a layout workbook cannot be parsed."""


@dataclass
class ConditionLayout:
    concentration: Decimal | None  # None for control
    wells: list[str]
    is_control: bool


@dataclass
class ExperimentLayout:
    date: date
    plate_wells: int
    conditions: list[ConditionLayout]


def parse_layout_xlsx(path: str | Path) -> ExperimentLayout:
    """
    Parse an Axion layout workbook (``*_LO.xlsx``).

    Sheet layout (first worksheet):
    - Column A contains keys (Date, Wells, Groups, then concentrations/Control).
    - Column B contains values (date, plate size, and well strings).
    - Conditions start after the ``Groups`` row.
    """
    path = Path(path)
    if not path.exists():
        raise LayoutError(f"Layout file does not exist: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb.worksheets[0]

    experiment_date: date | None = None
    plate_wells: int | None = None
    conditions: list[ConditionLayout] = []
    in_groups_section = False

    for key_cell, value_cell in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(key_cell).strip() if key_cell is not None else ""
        value = value_cell

        if not key:
            continue

        lowered = key.lower()
        if lowered == "date":
            experiment_date = _parse_date(value, source_path=path)
        elif lowered == "wells":
            plate_wells = _parse_plate_wells(value)
        elif lowered == "groups":
            in_groups_section = True
        elif in_groups_section:
            wells = parse_well_string(str(value).strip()) if value is not None else []
            if not wells:
                raise LayoutError(f"Condition '{key}' has no wells listed")

            if lowered == "control":
                condition = ConditionLayout(
                    concentration=None,
                    wells=wells,
                    is_control=True,
                )
            else:
                try:
                    concentration = Decimal(str(key))
                except Exception as exc:
                    raise LayoutError(f"Invalid concentration value '{key}'") from exc

                condition = ConditionLayout(
                    concentration=concentration,
                    wells=wells,
                    is_control=False,
                )
            conditions.append(condition)

    if experiment_date is None:
        raise LayoutError("Layout is missing a Date entry")
    if plate_wells is None:
        raise LayoutError("Layout is missing a Wells entry")
    if not conditions:
        raise LayoutError("Layout contains no conditions after the Groups row")
    if not any(cond.is_control for cond in conditions):
        raise LayoutError("Layout must include at least one Control group")

    _validate_wells(conditions, plate_wells)

    return ExperimentLayout(
        date=experiment_date,
        plate_wells=plate_wells,
        conditions=conditions,
    )


def _parse_date(value, *, source_path: Path | None = None) -> date:
    if value is None:
        raise LayoutError("Date cell is empty")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date (1900 epoch)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(value))).date()

    text = str(value).strip()
    parsed = _try_parse_date_text(text)
    if parsed is not None:
        return parsed

    if source_path is not None:
        fallback = _parse_date_from_filename(source_path)
        if fallback is not None:
            logger.warning(
                "Unrecognized layout date '%s' in %s; using date from filename %s",
                text,
                source_path,
                fallback.isoformat(),
            )
            return fallback

    raise LayoutError(f"Unrecognized date format '{value}'")


def _try_parse_date_text(text: str) -> date | None:
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _parse_date_from_filename(path: Path) -> date | None:
    stem = path.stem
    for token in re.split(r"[\s_]+", stem):
        if not token:
            continue
        if token.isdigit():
            if len(token) == 8:
                try:
                    return datetime.strptime(token, "%Y%m%d").date()
                except ValueError:
                    continue
            if len(token) == 6:
                try:
                    return datetime.strptime(token, "%y%m%d").date()
                except ValueError:
                    continue

        parsed = _try_parse_date_text(token)
        if parsed is not None:
            return parsed

    return None


def _parse_plate_wells(value) -> int:
    try:
        wells = int(value)
    except (TypeError, ValueError) as exc:  # noqa: PERF203
        raise LayoutError(f"Invalid plate well count '{value}'") from exc
    if wells <= 0:
        raise LayoutError("Plate well count must be positive")
    return wells


def _validate_wells(conditions: Iterable[ConditionLayout], plate_wells: int) -> None:
    seen: set[str] = set()
    dimensions = PLATE_DIMENSIONS.get(plate_wells)
    row_limit, col_limit = dimensions if dimensions else (None, None)

    for condition in conditions:
        for well in condition.wells:
            if well in seen:
                raise LayoutError(f"Duplicate well '{well}' in layout")
            seen.add(well)

            row_label = well[0].upper()
            try:
                col_number = int(well[1:])
            except ValueError as exc:
                raise LayoutError(f"Invalid well identifier '{well}'") from exc

            if col_number <= 0:
                raise LayoutError(f"Well number must be positive: '{well}'")
            if row_limit is not None:
                row_index = ord(row_label) - ord("A") + 1
                if row_index > row_limit:
                    raise LayoutError(
                        f"Well '{well}' exceeds row limit for {plate_wells}-well plate"
                    )
            if col_limit is not None and col_number > col_limit:
                raise LayoutError(
                    f"Well '{well}' exceeds column limit for {plate_wells}-well plate"
                )

    if dimensions:
        max_supported = dimensions[0] * dimensions[1]
        if len(seen) > max_supported:
            raise LayoutError(
                f"Layout lists {len(seen)} wells which exceeds plate capacity {max_supported}"
            )
    elif len(seen) > plate_wells:
        raise LayoutError(
            f"Layout lists {len(seen)} wells which exceeds plate capacity {plate_wells}"
        )